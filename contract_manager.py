# Contract Manager
# Author: OH.HALIM
# Version: 1.0.0
import tkinter as tk
from tkinter import ttk
from ttkbootstrap import Style, Window
from ttkbootstrap.constants import *
from ttkbootstrap.dialogs import Messagebox, Querybox
import tkinter.messagebox as messagebox
import tkinter.simpledialog as simpledialog
import pandas as pd
import datetime
from dateutil.relativedelta import relativedelta
import os
import re
class ContractManager:
    def __init__(self, root):
        self.root = root
        self.version = "1.0.0"
        self.author = "OH.HALIM"
        self.root.title(f"Contract Duration Manager v{self.version} - by {self.author}")
        self.root.geometry("1200x800")
        
        # Set theme colors - New modern color scheme
        self.colors = {
            'primary': '#2563EB',      # Blue-600
            'primary_light': '#60A5FA', # Blue-400
            'secondary': '#059669',     # Emerald-600
            'warning': '#D97706',       # Amber-600
            'danger': '#DC2626',        # Red-600
            'background': '#F8FAFC',    # Slate-50
            'surface': '#FFFFFF',       # White
            'text': '#0F172A',          # Slate-900
            'text_light': '#64748B',    # Slate-500
            'border': '#E2E8F0',        # Slate-200
            'hover': '#F1F5F9',         # Slate-100
        }
        
        # Initialize language (default to English)
        self.current_language = "fr"
        
        # Language dictionaries for English and French
        self.translations = {
            "en": {
                "title": "Contract Duration Manager",
                "header": "Contract Management System",
                "add_contract_frame": "Add New Contract",
                "prestataire": "Prestataire:",
                "contract_name": "Contract Name:",
                "operation": "Operation:",
                "fournisseur": "Fournisseur:",
                "start_date": "Start Date (DD-MM-YYYY):",
                "duration_months": "Duration (Months):",
                "add_button": "Add Contract",
                "save_button": "Save Contracts",
                "load_button": "Load Contracts",
                "delete_button": "Delete Selected",
                "extension_button": "Add Extension",
                "language_button": "Switch to French",
                "check_expired_button": "Check Expired",
                "column_prestataire": "Prestataire",
                "column_contract": "Contract",
                "column_operation": "Operation",
                "column_fournisseur": "Fournisseur",
                "column_start_date": "Start Date",
                "column_duration": "Duration",
                "column_expiration": "Expiration Date",
                "column_time_remaining": "Time Remaining",
                "column_status": "Status",
                "column_amendments": "Amendments",
                "status_active": "Active",
                "status_expiring": "Expiring Soon",
                "status_expired": "Expired",
                "time_year": "year",
                "time_years": "years",
                "time_month": "month",
                "time_months": "months",
                "time_day": "day",
                "time_days": "days",
                "msg_enter_name": "Please enter a contract name",
                "msg_invalid_date": "Invalid date format. Please use DD-MM-YYYY",
                "msg_save_success": "Contracts saved successfully to contracts.xlsx!",
                "msg_save_error": "Failed to save contracts",
                "msg_load_error": "Failed to load contracts",
                "msg_delete_warning": "Please select a contract to delete",
                "msg_delete_confirm": "Are you sure you want to delete the selected contract?",
                "msg_error": "Error",
                "msg_warning": "Warning",
                "msg_success": "Success",
                "msg_confirm": "Confirm",
                "msg_no_expired": "No expired contracts found.",
                "msg_expired_found": "Expired Contracts",
                "msg_expired_contracts": "The following contracts have expired:",
                "msg_select_contract": "Please select a contract to extend",
                "msg_extension_title": "Add Contract Extension",
                "msg_extension_prompt": "Enter extension duration in months:",
                "msg_extension_success": "Contract extended successfully",
                "msg_extension_added": "Extension added to contract",
                "btn_extend_selected": "Extend Selected Contract",
                "btn_extend_all": "Extend All Expired",
                "btn_close": "Close",
                "expired_by": "Expired by:",
                "avenant_number": "Amendment n°",
                "no_avenants": "No Amendments",
                "msg_unsaved_changes": "You have unsaved changes. Do you want to save them before closing?",
                "btn_yes": "Yes",
                "btn_no": "No",
                "btn_cancel": "Cancel",
                "search_placeholder": "Search by contract name or prestataire...",
                "search_button": "Search",
                "reset_button": "Reset",
                "filter_by": "Filter by:",
                "filter_all": "All",
                "filter_active": "Active",
                "filter_expiring": "Expiring Soon",
                "filter_expired": "Expired",
                "no_results": "No matching contracts found.",
                "search_results": "Search Results"
            },
            "fr": {
                "btn_yes": "Oui",
                "btn_no": "Non",
                "btn_cancel": "Annuler",
                "title": "Gestionnaire de Durée de Contrat",
                "header": "Système de Gestion des Contrats",
                "add_contract_frame": "Ajouter un Nouveau Contrat",
                "prestataire": "Prestataire:",
                "contract_name": "Nom du Contrat:",
                "operation": "Opération:",
                "fournisseur": "Fournisseur:",
                "start_date": "Date de Début (JJ-MM-AAAA):",
                "duration_months": "Durée (Mois):",
                "add_button": "Ajouter Contrat",
                "save_button": "Enregistrer",
                "load_button": "Charger",
                "delete_button": "Supprimer",
                "extension_button": "Ajouter Avenant",
                "language_button": "Passer en Anglais",
                "check_expired_button": "Vérifier Expirés",
                "column_prestataire": "Prestataire",
                "column_contract": "Contrat",
                "column_operation": "Opération",
                "column_fournisseur": "Fournisseur",
                "column_start_date": "Date de Début",
                "column_duration": "Durée",
                "column_expiration": "Date d'Expiration",
                "column_time_remaining": "Temps Restant",
                "column_status": "Statut",
                "column_amendments": "Avenants",
                "status_active": "Actif",
                "status_expiring": "Expiration Proche",
                "status_expired": "Expiré",
                "time_year": "an",
                "time_years": "ans",
                "time_month": "mois",
                "time_months": "mois",
                "time_day": "jour",
                "time_days": "jours",
                "msg_enter_name": "Veuillez entrer un nom de contrat",
                "msg_invalid_date": "Format de date invalide. Veuillez utiliser JJ-MM-AAAA",
                "msg_save_success": "Contrats enregistrés avec succès dans contracts.xlsx!",
                "msg_save_error": "Échec de l'enregistrement des contrats",
                "msg_load_error": "Échec du chargement des contrats",
                "msg_delete_warning": "Veuillez sélectionner un contrat à supprimer",
                "msg_delete_confirm": "Êtes-vous sûr de vouloir supprimer le contrat sélectionné?",
                "msg_error": "Erreur",
                "msg_warning": "Avertissement",
                "msg_success": "Succès",
                "msg_confirm": "Confirmer",
                "msg_no_expired": "Aucun contrat expiré trouvé.",
                "msg_expired_found": "Contrats Expirés",
                "msg_expired_contracts": "Les contrats suivants ont expiré:",
                "msg_select_contract": "Veuillez sélectionner un contrat à prolonger",
                "msg_extension_title": "Ajouter un Avenant",
                "msg_extension_prompt": "Entrez la durée de prolongation en mois:",
                "msg_extension_success": "Contrat prolongé avec succès",
                "msg_extension_added": "Avenant ajouté au contrat",
                "btn_extend_selected": "Prolonger le Contrat Sélectionné",
                "btn_extend_all": "Prolonger Tous les Expirés",
                "btn_close": "Fermer",
                "expired_by": "Expiré depuis:",
                "avenant_number": "Avenant n°",
                "no_avenants": "Aucun Avenant",
                "msg_unsaved_changes": "Vous avez des modifications non enregistrées. Voulez-vous les enregistrer avant de fermer ?",
                "search_placeholder": "Rechercher par nom de contrat ou prestataire...",
                "search_button": "Rechercher",
                "reset_button": "Réinitialiser",
                "filter_by": "Filtrer par:",
                "filter_all": "Tous",
                "filter_active": "Actifs",
                "filter_expiring": "Expiration Proche",
                "filter_expired": "Expirés",
                "no_results": "Aucun contrat correspondant trouvé.",
                "search_results": "Résultats de Recherche"
            }
        }
        
        # Configure modern styles
        self.style = root.style
        
        # Configure modern styles
        self.style.configure("Modern.TFrame", 
                           background=self.colors['background'])
        
        self.style.configure("Modern.TLabel", 
                           background=self.colors['background'],
                           foreground=self.colors['text'],
                           font=("Segoe UI", 10))
        
        self.style.configure("Header.TLabel", 
                           background=self.colors['background'],
                           foreground=self.colors['primary'],
                           font=("Segoe UI", 20, "bold"))
        
        self.style.configure("Modern.TButton", 
                           font=("Segoe UI", 10, "bold"),
                           padding=(15, 8),
                           borderwidth=0,
                           relief="flat",
                           background=self.colors['primary'],
                           foreground='white',
                           border_radius=8)
        
        self.style.map("Modern.TButton",
                      background=[('active', self.colors['primary_light']),
                                ('!active', self.colors['primary']),
                                ('hover', self.colors['primary_light'])],
                      foreground=[('active', 'white'),
                                ('!active', 'white'),
                                ('hover', 'white')])
        
        # Configure smaller button style for search/reset
        self.style.configure("Small.TButton", 
                           font=("Segoe UI", 9, "bold"),
                           padding=(8, 6),
                           borderwidth=0,
                           relief="flat",
                           background=self.colors['primary'],
                           foreground='white',
                           border_radius=6)
        
        self.style.map("Small.TButton",
                      background=[('active', self.colors['primary_light']),
                                ('!active', self.colors['primary']),
                                ('hover', self.colors['primary_light'])],
                      foreground=[('active', 'white'),
                                ('!active', 'white'),
                                ('hover', 'white')])
        
        # Configure Treeview style
        self.style.configure("Treeview",
                           background=self.colors['surface'],
                           foreground=self.colors['text'],
                           rowheight=35,
                           fieldbackground=self.colors['surface'],
                           font=("Segoe UI", 10),
                           borderwidth=0)
        
        self.style.configure("Treeview.Heading",
                           background=self.colors['primary'],
                           foreground='white',
                           font=("Segoe UI", 10, "bold"),
                           padding=8)
        
        self.style.map("Treeview.Heading",
                      background=[('active', self.colors['primary_light'])])
        
        # Create main frame with padding and rounded corners
        self.main_frame = ttk.Frame(self.root, style="Modern.TFrame", padding=30)
        self.main_frame.grid(row=0, column=0, sticky=(N, S, E, W))
        
        # Create header frame with modern styling
        header_frame = ttk.Frame(self.main_frame, style="Modern.TFrame")
        header_frame.grid(row=0, column=0, columnspan=2, sticky=(E, W), pady=(0, 30))
        
        # Add header label with modern styling
        self.header_label = ttk.Label(header_frame, 
                                    text=self.get_text("header"),
                                    style="Header.TLabel")
        self.header_label.grid(row=0, column=0, sticky=W)
        
        # Create a frame for the right-side buttons in the header
        header_buttons_frame = ttk.Frame(header_frame, style="Modern.TFrame")
        header_buttons_frame.grid(row=0, column=1, sticky=E)
        
        # Add language toggle button with modern styling
        self.language_button = ttk.Button(header_buttons_frame,
                                        text=self.get_text("language_button"),
                                        command=self.toggle_language,
                                        style="Modern.TButton")
        self.language_button.grid(row=0, column=0, padx=8)
        
        # Add check expired contracts button with modern styling
        self.check_expired_button = ttk.Button(header_buttons_frame,
                                             text=self.get_text("check_expired_button"),
                                             command=self.check_expired_contracts,
                                             style="Modern.TButton")
        self.check_expired_button.grid(row=0, column=1, padx=8)
        
        # Create input frame with modern styling
        self.input_frame = ttk.LabelFrame(self.main_frame,
                                        text=self.get_text("add_contract_frame"),
                                        style="Modern.TFrame",
                                        padding=20)
        self.input_frame.grid(row=1, column=0, columnspan=2, sticky=(E, W), pady=(0, 30))
        
        # Prestataire with modern styling
        self.prestataire_label = ttk.Label(self.input_frame,
                                         text=self.get_text("prestataire"),
                                         style="Modern.TLabel")
        self.prestataire_label.grid(row=0, column=0, sticky=W, padx=10, pady=10)
        self.prestataire = ttk.Entry(self.input_frame, width=30)
        self.prestataire.grid(row=0, column=1, padx=10, pady=10)
        
        # Contract Name with modern styling
        self.contract_name_label = ttk.Label(self.input_frame,
                                           text=self.get_text("contract_name"),
                                           style="Modern.TLabel")
        self.contract_name_label.grid(row=0, column=2, sticky=W, padx=10, pady=10)
        self.contract_name = ttk.Entry(self.input_frame, width=30)
        self.contract_name.grid(row=0, column=3, padx=10, pady=10)
        
        # Operation with modern styling
        self.operation_label = ttk.Label(self.input_frame,
                                       text=self.get_text("operation"),
                                       style="Modern.TLabel")
        self.operation_label.grid(row=1, column=0, sticky=W, padx=10, pady=10)
        self.operation = ttk.Entry(self.input_frame, width=30)
        self.operation.grid(row=1, column=1, padx=10, pady=10)
        
        # Fournisseur with modern styling
        self.fournisseur_label = ttk.Label(self.input_frame,
                                         text=self.get_text("fournisseur"),
                                         style="Modern.TLabel")
        self.fournisseur_label.grid(row=1, column=2, sticky=W, padx=10, pady=10)
        self.fournisseur = ttk.Entry(self.input_frame, width=30)
        self.fournisseur.grid(row=1, column=3, padx=10, pady=10)
        
        # Start Date with modern styling
        self.start_date_label = ttk.Label(self.input_frame,
                                        text=self.get_text("start_date"),
                                        style="Modern.TLabel")
        self.start_date_label.grid(row=2, column=0, sticky=W, padx=10, pady=10)
        self.start_date = ttk.Entry(self.input_frame, width=12)
        self.start_date.grid(row=2, column=1, padx=10, pady=10)
        self.start_date.insert(0, datetime.datetime.now().strftime("%d-%m-%Y"))
        
        # Duration with modern styling
        self.duration_label = ttk.Label(self.input_frame,
                                      text=self.get_text("duration_months"),
                                      style="Modern.TLabel")
        self.duration_label.grid(row=2, column=2, sticky=W, padx=10, pady=10)
        self.months = ttk.Spinbox(self.input_frame, from_=1, to=1200, width=5)
        self.months.grid(row=2, column=3, padx=10, pady=10)
        self.months.insert(0, "1")
        
        # Add Contract Button with modern styling
        self.add_button = ttk.Button(self.input_frame,
                                   text=self.get_text("add_button"),
                                   command=self.add_contract,
                                   style="Modern.TButton")
        self.add_button.grid(row=3, column=0, columnspan=4, padx=10, pady=10)
        
        # Create search and filter frame
        self.search_frame = ttk.Frame(self.main_frame, style="Modern.TFrame")
        self.search_frame.grid(row=2, column=0, columnspan=2, sticky=(E, W), pady=(0, 15))
        
        # Search entry
        self.search_var = tk.StringVar()
        self.search_entry = ttk.Entry(self.search_frame, width=40, textvariable=self.search_var)
        self.search_entry.grid(row=0, column=0, padx=(0, 10), pady=10)
        self.search_entry.insert(0, self.get_text("search_placeholder"))
        self.search_entry.bind("<FocusIn>", self.on_search_focus_in)
        self.search_entry.bind("<FocusOut>", self.on_search_focus_out)
        self.search_entry.bind("<Return>", lambda event: self.search_contracts())
        
        # Search button
        self.search_button = ttk.Button(self.search_frame,
                                      text=self.get_text("search_button"),
                                      command=self.search_contracts,
                                      style="Small.TButton")
        self.search_button.grid(row=0, column=1, padx=(0, 10), pady=10)
        
        # Reset search button
        self.reset_button = ttk.Button(self.search_frame,
                                     text=self.get_text("reset_button"),
                                     command=self.reset_search,
                                     style="Small.TButton")
        self.reset_button.grid(row=0, column=2, padx=(0, 20), pady=10)
        
        # Filter label
        self.filter_label = ttk.Label(self.search_frame,
                                    text=self.get_text("filter_by"),
                                    style="Modern.TLabel")
        self.filter_label.grid(row=0, column=3, padx=(0, 10), pady=10)
        
        # Filter combobox
        self.filter_var = tk.StringVar()
        self.filter_combo = ttk.Combobox(self.search_frame, 
                                       textvariable=self.filter_var, 
                                       width=15,
                                       state="readonly")
        self.filter_combo.grid(row=0, column=4, pady=10)
        
        # Update filter options based on current language
        self.update_filter_options()
        
        # Bind the combobox selection to filter contracts
        self.filter_combo.bind("<<ComboboxSelected>>", lambda event: self.filter_contracts())
        
        # Create a Frame to hold both treeview and status column
        self.display_frame = ttk.Frame(self.main_frame, style="Modern.TFrame")
        self.display_frame.grid(row=3, column=0, columnspan=2, sticky=(N, S, E, W))
        
        # Create a Treeview with all columns including Amendments column
        self.tree = ttk.Treeview(self.display_frame,
                                columns=("Prestataire", "Contract", "Operation", "Fournisseur",
                                       "Start Date", "Duration", "Expiration Date", 
                                       "Time Remaining", "Status", "Amendments"),
                                show="headings",
                                style="Treeview")
        
        # Set column headings based on language
        self.update_treeview_headings()
        
        # Set column widths
        self.tree.column("Prestataire", width=150)
        self.tree.column("Contract", width=150)
        self.tree.column("Operation", width=150)
        self.tree.column("Fournisseur", width=150)
        self.tree.column("Start Date", width=120)
        self.tree.column("Duration", width=120)
        self.tree.column("Expiration Date", width=120)
        self.tree.column("Time Remaining", width=140)
        self.tree.column("Status", width=120)
        self.tree.column("Amendments", width=120)
        
        # Add scrollbar for treeview with modern styling
        tree_scrollbar = ttk.Scrollbar(self.display_frame, orient=VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=tree_scrollbar.set)
        
        # Position tree and scrollbar
        self.tree.grid(row=0, column=0, sticky=(N, S, E, W))
        tree_scrollbar.grid(row=0, column=1, sticky=(N, S))
        
        # Make the tree take up most of the space
        self.display_frame.columnconfigure(0, weight=1)
        self.display_frame.rowconfigure(0, weight=1)
        
        # Configure tag styles for the status column
        self.tree.tag_configure('critical', background='#FEE2E2')  # Red-100
        self.tree.tag_configure('warning', background='#FEF3C7')   # Amber-100
        self.tree.tag_configure('active', background='#D1FAE5')    # Emerald-100
        self.tree.tag_configure('expired', background='#FEE2E2')   # Red-100
        
        # Buttons frame with modern styling
        self.button_frame = ttk.Frame(self.main_frame, style="Modern.TFrame")
        self.button_frame.grid(row=4, column=0, columnspan=2, pady=30)
        
        # Save and Load buttons with modern styling
        self.save_button = ttk.Button(self.button_frame,
                                    text=self.get_text("save_button"),
                                    command=self.save_to_excel,
                                    style="Modern.TButton")
        self.save_button.grid(row=0, column=0, padx=8)
        
        self.load_button = ttk.Button(self.button_frame,
                                    text=self.get_text("load_button"),
                                    command=self.load_from_excel,
                                    style="Modern.TButton")
        self.load_button.grid(row=0, column=1, padx=8)
        
        self.delete_button = ttk.Button(self.button_frame,
                                      text=self.get_text("delete_button"),
                                      command=self.delete_contract,
                                      style="Modern.TButton")
        self.delete_button.grid(row=0, column=2, padx=8)
        
        # Add extension button with modern styling
        self.extension_button = ttk.Button(self.button_frame,
                                         text=self.get_text("extension_button"),
                                         command=self.add_contract_extension,
                                         style="Modern.TButton")
        self.extension_button.grid(row=0, column=3, padx=8)
        
        # Footer frame for version and author info
        self.footer_frame = ttk.Frame(self.root, style="Modern.TFrame")
        self.footer_frame.grid(row=5, column=0, sticky=(E, W), pady=(0, 0))
        self.footer_label = ttk.Label(self.footer_frame,
                                    text=f"v{self.version}  |  by {self.author}",
                                    style="Modern.TLabel",
                                    font=("Segoe UI", 9, "italic"))
        self.footer_label.pack(anchor='center', pady=4)
        
        # Configure grid weights for main layout
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        self.main_frame.columnconfigure(0, weight=1)
        self.main_frame.rowconfigure(3, weight=1)  # Display frame should expand
        
        # Bind tree selection to update status displays
        self.tree.bind("<<TreeviewSelect>>", self.on_tree_select)
        
        # Initialize contracts list and saved items from treeview
        self.contracts = []
        self.all_treeview_items = []
        
        # Load existing contracts if any
        self.load_from_excel()
        
        # Initialize unsaved changes flag
        self.unsaved_changes = False
        
        # Check for expired contracts on startup
        self.root.after(1000, self.check_expired_contracts_silently)
    
    def get_text(self, key):
        """Get translated text for a key."""
        return self.translations[self.current_language].get(key, key)
    
    def toggle_language(self):
        """Toggle between English and French."""
        self.current_language = "fr" if self.current_language == "en" else "en"
        
        # Update UI elements with the new language
        self.root.title(self.get_text("title"))
        self.header_label.config(text=self.get_text("header"))
        self.language_button.config(text=self.get_text("language_button"))
        self.check_expired_button.config(text=self.get_text("check_expired_button"))
        self.input_frame.config(text=self.get_text("add_contract_frame"))
        self.prestataire_label.config(text=self.get_text("prestataire"))
        self.contract_name_label.config(text=self.get_text("contract_name"))
        self.operation_label.config(text=self.get_text("operation"))
        self.fournisseur_label.config(text=self.get_text("fournisseur"))
        self.start_date_label.config(text=self.get_text("start_date"))
        self.duration_label.config(text=self.get_text("duration_months"))
        self.add_button.config(text=self.get_text("add_button"))
        self.save_button.config(text=self.get_text("save_button"))
        self.load_button.config(text=self.get_text("load_button"))
        self.delete_button.config(text=self.get_text("delete_button"))
        self.extension_button.config(text=self.get_text("extension_button"))
        
        # Update search and filter UI
        self.update_search_placeholder()
        self.search_button.config(text=self.get_text("search_button"))
        self.reset_button.config(text=self.get_text("reset_button"))
        self.filter_label.config(text=self.get_text("filter_by"))
        self.update_filter_options()
        
        # Update treeview headers
        self.update_treeview_headings()
        
        # Refresh all items in the treeview to update status and duration text
        self.refresh_treeview_items()
    
    def update_search_placeholder(self):
        """Update search entry placeholder text based on current language."""
        current_text = self.search_var.get()
        placeholder = self.get_text("search_placeholder")
        # Only update if it's the placeholder text
        if current_text == self.translations["en"]["search_placeholder"] or current_text == self.translations["fr"]["search_placeholder"]:
            self.search_entry.delete(0, tk.END)
            self.search_entry.insert(0, placeholder)
    
    def on_search_focus_in(self, event):
        """Clear placeholder text when search entry gets focus."""
        if self.search_var.get() == self.get_text("search_placeholder"):
            self.search_entry.delete(0, tk.END)
    
    def on_search_focus_out(self, event):
        """Add placeholder text when search entry loses focus if empty."""
        if not self.search_var.get():
            self.search_entry.insert(0, self.get_text("search_placeholder"))
    
    def update_filter_options(self):
        """Update filter combobox options based on current language."""
        current_value = self.filter_var.get()
        
        # Update combobox values
        filter_options = [
            self.get_text("filter_all"),
            self.get_text("filter_active"),
            self.get_text("filter_expiring"),
            self.get_text("filter_expired")
        ]
        self.filter_combo['values'] = filter_options
        
        # Try to maintain the same selection after language change
        if current_value:
            # Find corresponding value in the new language
            if current_value == self.translations["en"]["filter_all"] or current_value == self.translations["fr"]["filter_all"]:
                self.filter_var.set(self.get_text("filter_all"))
            elif current_value == self.translations["en"]["filter_active"] or current_value == self.translations["fr"]["filter_active"]:
                self.filter_var.set(self.get_text("filter_active"))
            elif current_value == self.translations["en"]["filter_expiring"] or current_value == self.translations["fr"]["filter_expiring"]:
                self.filter_var.set(self.get_text("filter_expiring"))
            elif current_value == self.translations["en"]["filter_expired"] or current_value == self.translations["fr"]["filter_expired"]:
                self.filter_var.set(self.get_text("filter_expired"))
            else:
                # Default to "All" if no selection
                self.filter_var.set(self.get_text("filter_all"))
    
    def apply_status_color(self, item_id, status, tag):
        """Apply color tag to status cell."""
        # Apply tag to the item
        self.tree.item(item_id, tags=(tag,))
    
    def check_expired_contracts_silently(self):
        """Check for expired contracts without showing a dialog if none are found."""
        expired_contracts = []
        
        for item in self.tree.get_children():
            # Access values by column identifier and parse expiration date
            item_values = self.tree.item(item)['values']
            values = {
                self.tree.heading(col, 'text'): item_values[idx]
                for idx, col in enumerate(self.tree['columns'])
            }
            
            try:
                expiration_date_str = values[self.get_text("column_expiration")]
                expiration_date = datetime.datetime.strptime(expiration_date_str, "%d-%m-%Y").date()
                
                today = datetime.datetime.now().date()
                days_remaining = (expiration_date - today).days
                
                # Get the current status text from the treeview
                status_text_from_tree = values[self.get_text("column_status")]
                
                # If the contract is expired (days remaining is negative or status is explicitly expired)
                if days_remaining < 0 or status_text_from_tree in [self.get_text('status_expired'), self.translations['en']['status_expired']]:
                    # Access prestataire and contract name using column identifiers from the mapped values
                    prestataire_name = values[self.get_text("column_prestataire")]
                    contract_name = values[self.get_text("column_contract")]
                    expired_contracts.append((item, prestataire_name, contract_name, days_remaining))
            except (ValueError, KeyError) as e:
                print(f"Error processing item for expired check: {item_values} - {e}")
                # Continue to the next item if there's an error with this one
                continue
        
        # If expired contracts are found, show notification
        if expired_contracts:
            self.show_expired_dialog(expired_contracts)
    
    def check_expired_contracts(self):
        """Check for expired contracts and show a dialog with the results."""
        expired_contracts = []
        
        for item in self.tree.get_children():
            # Access values by column identifier and parse expiration date
            item_values = self.tree.item(item)['values']
            values = {
                self.tree.heading(col, 'text'): item_values[idx]
                for idx, col in enumerate(self.tree['columns'])
            }
            
            try:
                expiration_date_str = values[self.get_text("column_expiration")]
                expiration_date = datetime.datetime.strptime(expiration_date_str, "%d-%m-%Y").date()
                
                today = datetime.datetime.now().date()
                days_remaining = (expiration_date - today).days
                
                # Get the current status text from the treeview
                status_text_from_tree = values[self.get_text("column_status")]
                
                # If the contract is expired (days remaining is negative or status is explicitly expired)
                if days_remaining < 0 or status_text_from_tree in [self.get_text('status_expired'), self.translations['en']['status_expired']]:
                    # Access prestataire and contract name using column identifiers from the mapped values
                    prestataire_name = values[self.get_text("column_prestataire")]
                    contract_name = values[self.get_text("column_contract")]
                    expired_contracts.append((item, prestataire_name, contract_name, days_remaining))
            except (ValueError, KeyError) as e:
                print(f"Error processing item for expired check: {item_values} - {e}")
                # Continue to the next item if there's an error with this one
                continue
        
        # Show message based on results
        if not expired_contracts:
            messagebox.showinfo(self.get_text("msg_expired_found"), self.get_text("msg_no_expired"))
        else:
            # Create custom dialog with option to add extensions
            self.show_expired_dialog(expired_contracts)
    
    def show_expired_dialog(self, expired_contracts):
        """Show a custom dialog for expired contracts with extension options."""
        # Create a dialog window
        dialog = tk.Toplevel(self.root)
        dialog.title(self.get_text("msg_expired_found"))
        dialog.geometry("500x400")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Configure style for the dialog
        dialog.configure(background=self.colors['background'])
        
        # Create a frame to hold the content
        content_frame = tk.Frame(dialog, background=self.colors['background'])
        content_frame.pack(fill=tk.BOTH, expand=True)
        
        # Add a header label with modern styling
        tk.Label(content_frame,
                 text=self.get_text("msg_expired_contracts"),
                 font=("Segoe UI", 10, "bold"),
                 background=self.colors['primary'],
                 foreground='white').pack(anchor=tk.W, pady=(0, 10))
        
        # Create a frame for the list of expired contracts
        list_frame = tk.Frame(content_frame, background=self.colors['background'])
        list_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        # Create a listbox to display expired contracts
        listbox_frame = tk.Frame(list_frame, background=self.colors['surface'])
        listbox_frame.pack(fill=tk.BOTH, expand=True, side=tk.LEFT)
        
        # Create listbox with modern styling
        expired_list = tk.Listbox(listbox_frame,
                                width=50,
                                height=10,
                                font=("Segoe UI", 10),
                                bg=self.colors['surface'],
                                fg=self.colors['text'],
                                selectbackground=self.colors['primary'],
                                selectforeground='white',
                                activestyle='none',
                                relief='flat',
                                highlightthickness=1,
                                highlightbackground=self.colors['primary_light'])
        expired_list.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Add scrollbar with modern styling
        list_scrollbar = tk.Scrollbar(listbox_frame,
                                     orient=tk.VERTICAL,
                                     command=expired_list.yview)
        expired_list.configure(yscrollcommand=list_scrollbar.set)
        list_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Map Listbox indices to item_ids
        expired_item_ids = []
        
        # Add expired contracts to the listbox
        for i, (item_id, name, contract_name, days) in enumerate(expired_contracts):
            time_desc = self.format_time_remaining(days)
            expired_list.insert(tk.END, f"{name} - {contract_name} ({self.get_text('expired_by')} {time_desc[1:]})")
            expired_item_ids.append(item_id)
        
        # Select the first item by default
        if expired_contracts:
            expired_list.selection_set(0)
        
        # Create a frame for buttons with modern styling
        button_frame = tk.Frame(content_frame, background=self.colors['background'])
        button_frame.pack(fill=tk.X, pady=(10, 0))
        
        # Add 'Ajouter Avenant' button for the selected contract
        tk.Button(button_frame,
                  text=self.get_text("extension_button"),
                  command=lambda: self.extend_selected_contract(dialog, expired_list, expired_item_ids),
                  font=("Segoe UI", 10, "bold"),
                  background=self.colors['primary'],
                  foreground='white',
                  padx=5,
                  pady=5).pack(side=tk.LEFT)
        
        # Keep only the 'Fermer' (Close) button
        tk.Button(button_frame,
                  text=self.get_text("btn_close"),
                  command=dialog.destroy,
                  font=("Segoe UI", 10, "bold"),
                  background=self.colors['primary'],
                  foreground='white',
                  padx=5,
                  pady=5).pack(side=tk.RIGHT)
    
    def extend_selected_contract(self, dialog, expired_list, expired_item_ids):
        """Extend the selected contract from the expired contracts list."""
        # Get the selected item
        selection = expired_list.curselection()
        if not selection:
            messagebox.showwarning(self.get_text("msg_warning"), self.get_text("msg_select_contract"))
            return
        
        # Get the item ID from the mapping
        idx = selection[0]
        item_id = expired_item_ids[idx]
        
        # Close the dialog
        dialog.destroy()
        
        # Add an extension to the contract
        self.extend_contract(item_id)
        
        # After extending, refresh the expired contracts check
        self.check_expired_contracts_silently()
    
    def show_extension_dialog(self, title, prompt, minvalue=1, maxvalue=1200, default=1):
        """Show a modern modal dialog to enter extension duration in months."""
        result = {'value': None}
        dialog = tk.Toplevel(self.root)
        dialog.title(title)
        dialog.geometry("350x160")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.configure(background=self.colors['background'])
        # Center the dialog
        dialog.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() // 2) - (350 // 2)
        y = self.root.winfo_y() + (self.root.winfo_height() // 2) - (160 // 2)
        dialog.geometry(f"350x160+{x}+{y}")
        # Prompt label
        label = tk.Label(dialog, text=prompt, font=("Segoe UI", 10), background=self.colors['background'], foreground=self.colors['text'])
        label.pack(pady=(18, 8))
        # Spinbox for months
        spin_var = tk.IntVar(value=default)
        spinbox = tk.Spinbox(dialog, from_=minvalue, to=maxvalue, textvariable=spin_var, width=8, font=("Segoe UI", 10))
        spinbox.pack(pady=(0, 12))
        # Button frame
        btn_frame = tk.Frame(dialog, background=self.colors['background'])
        btn_frame.pack(pady=(0, 8))
        def on_ok():
            try:
                val = int(spinbox.get())
                if minvalue <= val <= maxvalue:
                    result['value'] = val
                    dialog.destroy()
            except Exception:
                pass
        def on_cancel():
            dialog.destroy()
        ok_btn = tk.Button(btn_frame, text="OK", font=("Segoe UI", 10, "bold"), command=on_ok, background=self.colors['primary'], foreground='white', padx=8, pady=5)
        ok_btn.pack(side=tk.LEFT)
        cancel_btn = tk.Button(btn_frame, text="Cancel", font=("Segoe UI", 10, "bold"), command=on_cancel, background=self.colors['primary'], foreground='white', padx=8, pady=5)
        cancel_btn.pack(side=tk.LEFT)
        dialog.wait_window()
        return result['value']

    def extend_contract(self, item_id, duration=None):
        """Add an extension to a contract (duration in months)."""
        # Get the contract details
        item_values = self.tree.item(item_id)['values']
        # Map values to dictionary for easier access by column name
        values = {
            self.tree.heading(col, 'text'): item_values[idx]
            for idx, col in enumerate(self.tree['columns'])
        }
        
        if not duration:
            # Use the modern dialog instead of simpledialog.askinteger
            duration = self.show_extension_dialog(
                self.get_text("msg_extension_title"),
                self.get_text("msg_extension_prompt"),
                minvalue=1, maxvalue=1200, default=1
            )
        
        if duration:
            # Parse the expiration date string from the treeview to a date object
            expiration_date_str = values[self.get_text("column_expiration")]
            expiration_date_obj = datetime.datetime.strptime(expiration_date_str, "%d-%m-%Y")
            
            # Calculate new expiration date (add months)
            new_expiration_date_obj = expiration_date_obj + relativedelta(months=duration)
            
            # Calculate days remaining based on the new expiration date object
            today = datetime.datetime.now().date()
            days_remaining = (new_expiration_date_obj.date() - today).days
            
            # Calculate and format time remaining based on new days_remaining
            time_remaining_str = self.format_time_remaining(days_remaining)
            
            # Determine status text and tag based on new days remaining
            if days_remaining < 0:
                status_text = self.get_text('status_expired')
                tag = 'expired'
            elif days_remaining < 35:
                status_text = self.get_text('status_expiring')
                tag = 'critical'
            elif days_remaining <= 70:
                status_text = self.get_text('status_expiring')
                tag = 'warning'
            else:
                status_text = self.get_text('status_active')
                tag = 'active'
            
            # Get original duration string from the treeview
            original_duration_str = values[self.get_text("column_duration")]
            
            # Format new duration string
            if duration == 1:
                extension_text = f"1 {self.get_text('time_month')}"
            else:
                extension_text = f"{duration} {self.get_text('time_months')}"
            
            # Get original status to check if the contract was expired
            original_status = values[self.get_text("column_status")]
            
            # Determine how to display the new duration based on original status
            if original_status in [self.get_text('status_expired'), self.translations['en']['status_expired']]:
                # If expired, show only the amendment duration
                new_duration_str = extension_text
            else:
                # If not expired, show cumulative duration
                new_duration_str = f"{original_duration_str} + {extension_text}"
            
            # Get current amendment status string and update
            current_amendment_str = values[self.get_text("column_amendments")]
            
            if current_amendment_str == self.get_text("no_avenants") or current_amendment_str == self.translations["en"]["no_avenants"]:
                # First amendment
                new_amendment_str = f"{self.get_text('avenant_number')}1"
            else:
                # Extract current amendment number and increment
                match = re.search(r'#(\d+)$|n°(\d+)$|\d+$', str(current_amendment_str))
                if match:
                    current_num = int(match.group(1) or match.group(2) or match.group(0))
                    new_amendment_str = f"{self.get_text('avenant_number')}{current_num + 1}"
                else:
                    # Fallback if regex fails or format is unexpected
                    new_amendment_str = str(current_amendment_str) # Keep the original if format is unknown
            
            # Create updated item values tuple
            new_item_values = (
                values[self.get_text("column_prestataire")],  # Prestataire (string)
                values[self.get_text("column_contract")],  # Contract name (string)
                values[self.get_text("column_operation")], # Operation (string)
                values[self.get_text("column_fournisseur")], # Fournisseur (string)
                values[self.get_text("column_start_date")], # Start Date (string)
                new_duration_str,  # Updated duration (string)
                new_expiration_date_obj.strftime("%d-%m-%Y"),  # New expiration date (formatted string)
                time_remaining_str, # Updated time remaining (string)
                status_text,  # Updated status (string)
                new_amendment_str  # Updated amendment (string)
            )
            
            # Update the item in the treeview
            self.tree.item(item_id, values=new_item_values, tags=(tag,))
            
            # Update the item in all_treeview_items for search/filter functionality
            for i, item in enumerate(self.all_treeview_items):
                if (item[0] == values[self.get_text("column_prestataire")] and 
                    item[1] == values[self.get_text("column_contract")]):
                    self.all_treeview_items[i] = new_item_values
                    break
            
            # Mark that there are unsaved changes
            self.unsaved_changes = True
            
            # Show success message
            messagebox.showinfo(self.get_text("msg_success"), 
                               f"{self.get_text('msg_extension_added')}: {values[self.get_text('column_contract')]}")
    
    def add_contract_extension(self):
        """Add an extension to the selected contract."""
        # Get the selected item
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning(self.get_text("msg_warning"), self.get_text("msg_select_contract"))
            return
        
        # Extend the contract
        self.extend_contract(selection[0])
        
        # After extending, refresh the expired contracts check
        self.check_expired_contracts_silently()
    
    def save_to_excel(self):
        """Save contracts to Excel file."""
        try:
            data = []
            # Use all_treeview_items instead of directly accessing the treeview
            for item_values in self.all_treeview_items:
                # Parse expiration date to calculate days remaining
                expiration_date_str = item_values[6]  # Expiration date is at index 6
                expiration_date = datetime.datetime.strptime(expiration_date_str, "%d-%m-%Y").date()
                
                today = datetime.datetime.now().date()
                days_remaining = (expiration_date - today).days
                
                data.append({
                    'Prestataire': item_values[0],
                    'Contract Name': item_values[1],
                    'Operation': item_values[2],
                    'Fournisseur': item_values[3],
                    'Start Date': item_values[4],
                    'Duration': item_values[5],
                    'Expiration Date': item_values[6],
                    'Days Remaining': days_remaining,  # Store raw number for calculations
                    'Time Remaining': item_values[7],  # Store formatted string for display
                    'Status': item_values[8],
                    'Amendments': item_values[9]   # Store amendments info
                })
            
            df = pd.DataFrame(data)
            
            # Create Excel writer object
            with pd.ExcelWriter('contracts.xlsx', engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Contracts')
                
                # Get the workbook and the worksheet
                workbook = writer.book
                worksheet = writer.sheets['Contracts']
                
                # Adjust column widths
                for idx, col in enumerate(df.columns):
                    max_length = max(
                        df[col].astype(str).apply(len).max(),
                        len(col)
                    )
                    worksheet.column_dimensions[chr(65 + idx)].width = max_length + 2
                
                # Add some basic formatting
                for row in worksheet.iter_rows(min_row=1, max_row=1):
                    for cell in row:
                        cell.font = cell.font.copy(bold=True)
                
                # Add conditional formatting for status column
                from openpyxl.styles import PatternFill
                from openpyxl.formatting.rule import CellIsRule
                
                # Create fills for different statuses
                red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
                yellow_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
                green_fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
                
                # Find the status column index
                status_col_idx = df.columns.get_loc('Status')
                status_col = chr(65 + status_col_idx)
                
                # Apply conditional formatting to status column
                # For English status text
                worksheet.conditional_formatting.add(f'{status_col}2:{status_col}{len(df)+1}',
                                                  CellIsRule(operator='equal', 
                                                           formula=["\"Expiring Soon\""], 
                                                           fill=red_fill))
                worksheet.conditional_formatting.add(f'{status_col}2:{status_col}{len(df)+1}',
                                                  CellIsRule(operator='equal', 
                                                           formula=["\"Expired\""], 
                                                           fill=red_fill))
                worksheet.conditional_formatting.add(f'{status_col}2:{status_col}{len(df)+1}',
                                                  CellIsRule(operator='equal', 
                                                           formula=["\"Active\""], 
                                                           fill=green_fill))
                
                # For French status text
                worksheet.conditional_formatting.add(f'{status_col}2:{status_col}{len(df)+1}',
                                                  CellIsRule(operator='equal', 
                                                           formula=["\"Expiration Proche\""], 
                                                           fill=red_fill))
                worksheet.conditional_formatting.add(f'{status_col}2:{status_col}{len(df)+1}',
                                                  CellIsRule(operator='equal', 
                                                           formula=["\"Expiré\""], 
                                                           fill=red_fill))
                worksheet.conditional_formatting.add(f'{status_col}2:{status_col}{len(df)+1}',
                                                  CellIsRule(operator='equal', 
                                                           formula=["\"Actif\""], 
                                                           fill=green_fill))
            
            messagebox.showinfo(self.get_text("msg_success"), self.get_text("msg_save_success"))
            # Reset unsaved changes flag after saving
            self.unsaved_changes = False
        except Exception as e:
            messagebox.showerror(self.get_text("msg_error"), f"{self.get_text('msg_save_error')}: {str(e)}")
    
    def load_from_excel(self):
        """Load contracts from Excel file."""
        try:
            if os.path.exists('contracts.xlsx'):
                df = pd.read_excel('contracts.xlsx')
                
                # Check if language is stored in the file and update if necessary
                if 'Language' in df.columns and not df.empty:
                    saved_language = df.iloc[0]['Language']
                    if saved_language in ('en', 'fr') and saved_language != self.current_language:
                        self.current_language = saved_language
                
                # Clear existing items
                for item in self.tree.get_children():
                    self.tree.delete(item)
                
                # Clear all_treeview_items
                self.all_treeview_items = []
                
                # Add items from Excel
                for _, row in df.iterrows():
                    # Get days remaining 
                    days_remaining = int(row['Days Remaining'])
                    
                    # Determine status based on days remaining
                    if days_remaining < 0:
                        status = self.get_text('status_expired')
                        tag = 'expired'
                    elif days_remaining < 35:
                        status = self.get_text('status_expiring')
                        tag = 'critical'
                    elif days_remaining <= 70:
                        status = self.get_text('status_expiring')
                        tag = 'warning'
                    else:
                        status = self.get_text('status_active')
                        tag = 'active'
                    
                    # Format time remaining
                    time_remaining = self.format_time_remaining(days_remaining)
                    
                    # Handle amendment information
                    amendments = row['Amendments']
                    if amendments == self.translations["en"]["no_avenants"] or amendments == self.translations["fr"]["no_avenants"]:
                         amendments_text = self.get_text("no_avenants")
                    elif "Amendment #" in str(amendments) or "Avenant n°" in str(amendments) or re.match(r'^\d+$', str(amendments)):
                        # Extract the amendment number
                        match = re.search(r'#(\d+)$|n°(\d+)$|\d+$', str(amendments))
                        if match:
                            num = match.group(1) or match.group(2) or match.group(0)
                            amendments_text = f"{self.get_text('avenant_number')}{num}"
                        else:
                            # Fallback if regex fails
                            amendments_text = str(amendments)
                    else:
                         # If not in expected format, keep as is
                         amendments_text = str(amendments)
                    
                    # Create a tuple of values in the correct order
                    item_values = (
                        row['Prestataire'],
                        row['Contract Name'],
                        row['Operation'],
                        row['Fournisseur'],
                        row['Start Date'],
                        row['Duration'],
                        row['Expiration Date'],
                        time_remaining,
                        status,
                        amendments_text
                    )
                    
                    # Add to all_treeview_items for search/filter functionality
                    self.all_treeview_items.append(item_values)
                    
                    # Add to treeview with appropriate tag
                    item_id = self.tree.insert("", "end", values=item_values, tags=(tag,))
                
                # Update search placeholder based on current language
                self.update_search_placeholder()
                
                # Update filter options based on current language
                self.update_filter_options()
                
        except FileNotFoundError:
            # No contracts.xlsx found, start with an empty treeview.
            pass
        except Exception as e:
            messagebox.showerror(self.get_text("msg_error"), f"{self.get_text('msg_load_error')}: {str(e)}")
    
    def delete_contract(self):
        """Delete selected contract."""
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning(self.get_text("msg_warning"), self.get_text("msg_delete_warning"))
            return
            
        if messagebox.askyesno(self.get_text("msg_confirm"), self.get_text("msg_delete_confirm")):
            for item in selected:
                # Get the item values
                item_values = self.tree.item(item)['values']
                
                # Remove from all_treeview_items
                for i, stored_item in enumerate(self.all_treeview_items):
                    if (stored_item[0] == item_values[0] and  # Prestataire
                        stored_item[1] == item_values[1]):    # Contract name
                        del self.all_treeview_items[i]
                        break
                
                # Delete from treeview
                self.tree.delete(item)
            
            # Mark that there are unsaved changes
            self.unsaved_changes = True
    
    def on_closing(self):
        """Handle window closing event, prompt to save if there are unsaved changes."""
        if self.unsaved_changes:
            result = self.show_unsaved_changes_dialog()
            
            if result is True:
                self.save_to_excel()
                self.root.destroy()
            elif result is False:  # User chose "No", discard changes and close
                self.root.destroy()
        else:
            self.root.destroy()

    def show_unsaved_changes_dialog(self):
        """Show a custom modal dialog to ask the user about saving unsaved changes."""
        result = {'value': None}  # Use None for Cancel, True for Yes, False for No
        dialog = tk.Toplevel(self.root)
        dialog.title(self.get_text("msg_warning"))
        dialog.geometry("600x150")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.configure(background=self.colors['background'])
        # Center the dialog
        dialog.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() // 2) - (600 // 2)
        y = self.root.winfo_y() + (self.root.winfo_height() // 2) - (150 // 2)
        dialog.geometry(f"600x150+{x}+{y}")
        # Message label
        label = tk.Label(dialog, text=self.get_text("msg_unsaved_changes"), font=("Segoe UI", 10), background=self.colors['background'], foreground=self.colors['text'])
        label.pack(pady=(20, 10))
        # Button frame
        btn_frame = tk.Frame(dialog, background=self.colors['background'])
        btn_frame.pack(pady=(0, 10))
        def on_yes():
            result['value'] = True
            dialog.destroy()
        def on_no():
            result['value'] = False
            dialog.destroy()
        def on_cancel():
            result['value'] = None
            dialog.destroy()
        # Yes button
        yes_btn = tk.Button(btn_frame, text=self.get_text("btn_yes"), font=("Segoe UI", 10, "bold"), command=on_yes, background=self.colors['primary'], foreground='white', padx=8, pady=5)
        yes_btn.pack(side=tk.LEFT, padx=5)
        # No button
        no_btn = tk.Button(btn_frame, text=self.get_text("btn_no"), font=("Segoe UI", 10, "bold"), command=on_no, background=self.colors['primary'], foreground='white', padx=8, pady=5)
        no_btn.pack(side=tk.LEFT, padx=5)
        # Cancel button
        cancel_btn = tk.Button(btn_frame, text=self.get_text("btn_cancel"), font=("Segoe UI", 10, "bold"), command=on_cancel, background=self.colors['primary'], foreground='white', padx=8, pady=5)
        cancel_btn.pack(side=tk.LEFT, padx=5)
        dialog.wait_window()
        return result['value']

    def add_contract(self):
        """Add a new contract to the list and treeview."""
        # Implement your logic here, or use your previous code for adding a contract
        pass

    def search_contracts(self):
        """Search contracts by contract name or prestataire."""
        query = self.search_var.get().strip().lower()
        if not query or query == self.get_text("search_placeholder").lower():
            self.reset_search()
            return

        # Clear the treeview
        for item in self.tree.get_children():
            self.tree.delete(item)

        # Filter and display matching contracts
        found = False
        for item_values in self.all_treeview_items:
            contract_name = str(item_values[1]).lower()
            prestataire = str(item_values[0]).lower()
            if query in contract_name or query in prestataire:
                tag = 'active'
                status = item_values[8]
                if status == self.get_text('status_expired'):
                    tag = 'expired'
                elif status == self.get_text('status_expiring'):
                    tag = 'critical'
                elif status == self.get_text('status_active'):
                    tag = 'active'
                elif status == self.get_text('status_expiring'):
                    tag = 'warning'
                self.tree.insert("", "end", values=item_values, tags=(tag,))
                found = True

        if not found:
            messagebox.showinfo(self.get_text("search_results"), self.get_text("no_results"))

    def reset_search(self):
        """Reset the search and show all contracts."""
        self.search_var.set("")
        self.update_search_placeholder()
        # Clear the treeview
        for item in self.tree.get_children():
            self.tree.delete(item)
        # Re-add all items
        for item_values in self.all_treeview_items:
            tag = 'active'
            status = item_values[8]
            if status == self.get_text('status_expired'):
                tag = 'expired'
            elif status == self.get_text('status_expiring'):
                tag = 'critical'
            elif status == self.get_text('status_active'):
                tag = 'active'
            elif status == self.get_text('status_expiring'):
                tag = 'warning'
            self.tree.insert("", "end", values=item_values, tags=(tag,))

    def update_treeview_headings(self):
        """Update treeview headings based on the current language."""
        self.tree.heading("Prestataire", text=self.get_text("column_prestataire"))
        self.tree.heading("Contract", text=self.get_text("column_contract"))
        self.tree.heading("Operation", text=self.get_text("column_operation"))
        self.tree.heading("Fournisseur", text=self.get_text("column_fournisseur"))
        self.tree.heading("Start Date", text=self.get_text("column_start_date"))
        self.tree.heading("Duration", text=self.get_text("column_duration"))
        self.tree.heading("Expiration Date", text=self.get_text("column_expiration"))
        self.tree.heading("Time Remaining", text=self.get_text("column_time_remaining"))
        self.tree.heading("Status", text=self.get_text("column_status"))
        self.tree.heading("Amendments", text=self.get_text("column_amendments"))

    def on_tree_select(self, event):
        """Handle tree selection event (currently does nothing)."""
        pass

    def format_time_remaining(self, days_remaining):
        """Format the time remaining as a human-readable string."""
        if days_remaining < 0:
            days_remaining = abs(days_remaining)
            prefix = "-"
        else:
            prefix = ""
        years, rem = divmod(days_remaining, 365)
        months, days = divmod(rem, 30)
        parts = []
        if years > 0:
            parts.append(f"{years} {self.get_text('time_years') if years > 1 else self.get_text('time_year')}")
        if months > 0:
            parts.append(f"{months} {self.get_text('time_months') if months > 1 else self.get_text('time_month')}")
        if days > 0 or not parts:
            parts.append(f"{days} {self.get_text('time_days') if days > 1 else self.get_text('time_day')}")
        return prefix + ", ".join(parts)

if __name__ == "__main__":
    root = Window(themename="cosmo")  # Use ttkbootstrap's Window with a modern theme
    app = ContractManager(root)
    root.protocol("WM_DELETE_WINDOW", app.on_closing)
    root.mainloop()
    

